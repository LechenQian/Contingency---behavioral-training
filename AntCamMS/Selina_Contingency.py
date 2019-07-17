'''
Created on July 17, 2019

@author: Hao Wu, modified by Selina Qian
'''
from ScopeFoundry import Measurement
from ScopeFoundry.measurement import MeasurementQThread
from ScopeFoundry.helper_funcs import sibling_path, load_qt_ui_file
from ScopeFoundry import h5_io
import pyqtgraph as pg
import numpy as np
import random

from scipy import ndimage
import time
from numpy.random import rand
# import PySpin
from qtpy import QtCore
from qtpy.QtCore import QObject
import os
import queue
from AntCamHW.daq_do.daq_do_dev import DAQSimpleDOTask
from AntCamHW.daq_di.daq_di_dev import DAQSimpleDITask

from openpyxl import Workbook


class SubMeasurementQThread(MeasurementQThread):
    '''
    Sub-Thread for different loops in measurement
    '''

    def __init__(self, run_func, parent=None):
        '''
        run_func: user-defined function to run in the loop
        parent = parent thread, usually None
        '''
        super(MeasurementQThread, self).__init__(parent)
        self.run_func = run_func
        self.interrupted = False

    def run(self):
        while not self.interrupted:
            self.run_func()
            if self.interrupted:
                break

    @QtCore.Slot()
    def interrupt(self):
        self.interrupted = True


class SelinaTraining(Measurement):
    # this is the name of the measurement that ScopeFoundry uses
    # when displaying your measurement and saving data related to it
    name = "selina_training"
    interrupt_subthread = QtCore.Signal(())

    def setup(self):
        """
        Runs once during App initialization.
        This is the place to load a user interface file,
        define settings, and set up data structures.
        """

        # Define ui file to be used as a graphical interface
        # This file can be edited graphically with Qt Creator
        # sibling_path function allows python to find a file in the same folder
        # as this python module
        self.ui_filename = sibling_path(__file__, "ant_watch_plot.ui")

        # Load ui file and convert it to a live QWidget of the user interface
        self.ui = load_qt_ui_file(self.ui_filename)

        # Measurement Specific Settings
        # This setting allows the option to save data to an h5 data file during a run
        # All settings are automatically added to the Microscope user interface
        self.settings.New('save_h5', dtype=bool, initial=True)
        self.settings.New('save_video', dtype=bool, initial=False)

        # x and y is for transmitting signal
        self.settings.New('x', dtype=float, initial=32, ro=True, vmin=0, vmax=63.5)
        self.settings.New('y', dtype=float, initial=32, ro=True, vmin=0, vmax=63.5)

        # added by Nune
        self.settings.New('filename', dtype=str, initial='trial')
        self.settings.New('in_trial', dtype=bool, initial=False)
        self.settings.New('view_only', dtype=bool, initial=False)
        self.settings.New('lick_status', dtype=int, initial=0)
        self.settings.New('play_frequency', dtype=int, initial=0)

        # Define how often to update display during a run
        self.display_update_period = 0.01

        # Convenient reference to the hardware used in the measurement
        self.wide_cam = self.app.hardware['wide_cam']
        self.recorder = self.app.hardware['flirrec']

        # setup experiment condition
        self.wide_cam.settings.frame_rate.update_value(8)
        self.wide_cam.read_from_hardware()



    def setup_figure(self):
        """
        Runs once during App initialization, after setup()
        This is the place to make all graphical interface initializations,
        build plots, etc.
        """
        # connect ui widgets to measurement/hardware settings or functions
        self.ui.start_pushButton.clicked.connect(self.start)
        self.ui.interrupt_pushButton.clicked.connect(self.interrupt)

        # Set up pyqtgraph graph_layout in the UI
        self.wide_cam_layout = pg.GraphicsLayoutWidget()
        self.ui.wide_cam_groupBox.layout().addWidget(self.wide_cam_layout)

        # create camera image graphs
        self.wide_cam_view = pg.ViewBox()
        self.wide_cam_layout.addItem(self.wide_cam_view)
        self.wide_cam_image = pg.ImageItem()
        self.wide_cam_view.addItem(self.wide_cam_image)

        # counter used for reducing refresh rate
        self.wide_disp_counter = 0

    def update_display(self):
        """
        Displays (plots) the numpy array self.buffer.
        This function runs repeatedly and automatically during the measurement run.
        its update frequency is defined by self.display_update_period
        """

        # check availability of display queue of the wide camera
        if not hasattr(self, 'wide_disp_queue'):
            pass
        elif self.wide_disp_queue.empty():
            pass
        else:
            try:
                wide_disp_image = self.wide_disp_queue.get()
                if type(wide_disp_image) == np.ndarray:
                    if wide_disp_image.shape == (
                    self.wide_cam.settings.height.value(), self.wide_cam.settings.width.value()):
                        try:
                            self.wide_cam_image.setImage(wide_disp_image)
                        except Exception as ex:
                            print('Error: %s' % ex)
            except Exception as ex:
                print("Error: %s" % ex)

    def run(self):

        """
        Runs when measurement is started. Runs in a separate thread from GUI.
        It should not update the graphical interface directly, and should only
        focus on data acquisition.
        """
        #         # first, create a data file
        #         if self.settings['save_h5']:
        #             # if enabled will create an HDF5 file with the plotted data
        #             # first we create an H5 file (by default autosaved to app.settings['save_dir']
        #             # This stores all the hardware and app meta-data in the H5 file
        #             self.h5file = h5_io.h5_base_file(app=self.app, measurement=self)
        #
        #             # create a measurement H5 group (folder) within self.h5file
        #             # This stores all the measurement meta-data in this group
        #             self.h5_group = h5_io.h5_create_measurement_group(measurement=self, h5group=self.h5file)
        #
        #             # create an h5 dataset to store the data
        #             self.buffer_h5 = self.h5_group.create_dataset(name  = 'buffer',
        #                                                           shape = self.buffer.shape,
        #                                                           dtype = self.buffer.dtype)

        # We use a try/finally block, so that if anything goes wrong during a measurement,
        # the finally block can clean things up, e.g. close the data file object.

        # self.wide_cam._dev.set_buffer_count(500)

        # if self.settings.save_video.value():
        #     save_dir = self.app.settings.save_dir.value()
        #     data_path = os.path.join(save_dir, self.app.settings.sample.value())
        #     try:
        #         os.makedirs(data_path)
        #     except OSError:
        #         print('directory already exists, writing to existing directory')
        #
        #     self.recorder.settings.path.update_value(data_path)

            # frame_rate = self.wide_cam.settings.frame_rate.value()
            # self.recorder.create_file('wide_mov',frame_rate)
        #
        #         # create a subthread and connect it to the interrupt subthread signal
        #         self.wide_disp_queue = queue.Queue(1000)
        #         self.camera_thread = SubMeasurementQThread(self.camera_action)
        #         self.interrupt_subthread.connect(self.camera_thread.interrupt)
        #         #start camera
        #         self.wide_cam.start()
        #
        #         #start camera subthread
        #         self.camera_thread.start()


        odors_cue = OdorGen([0, 1, 2, 3])
        odors_cue.assign_odor()
        self.reward_odor = odors_cue.set_rewardodor(index=0)
        odors_cue.initiate()
        # odors_cue.odors_DAQ[i]
        self.events_filename = '2019-7-17-test.xlsx'



        self.waterR = DAQSimpleDOTask('Dev1/port0/line0')
        self.waterR.low()
        # self.OdorOnCopy = DAQSimpleDOTask('Dev3/port2/line5')
        # self.OdorOnCopy.low()
        self.lickR = DAQSimpleDITask('Dev2_SELECT/port1/line0')

        # EVENT CODES
        # video recording start / start trial = 101
        # lick on = 11, lick off = 10
        # right
        # airpuff on = 81, off = 80

        # contingency reward odor on = 131, off = 130, water on = 51, right water off = 50
        # contingency no reward odor on = 141, off = 140, water on = 61, right water off = 60
        # non-contingency reward odor on = 151, off = 150, water on = 71, right water off = 70
        # non-contingency no reward odor on = 161, off = 160, water on = 81, right water off = 80

        # create excel workbook
        self.wb = Workbook()
        self.ws = self.wb.active


        #generate trial type
        numtrials = 200
        p_cont_noncont = 0.5
        p_USwCS = 0.5
        p_USwoCS = 0.5
        cont_reward = np.zeros(int(numtrials * p_cont_noncont * p_USwCS))  # code 0
        cont_noreward = np.ones(int(numtrials * p_cont_noncont * (1 - p_USwCS)))  # code 1
        temp_comb1 = np.concatenate((cont_reward, cont_noreward))

        noncont_reward = np.ones(int(numtrials * (1 - p_cont_noncont) * p_USwoCS)) * 2  # code 2
        noncont_noreward = np.ones(int(numtrials * (1 - p_cont_noncont) * (1 - p_USwoCS))) * 3  # code 3
        temp_comb2 = np.concatenate((noncont_noreward, noncont_reward))

        trialtypes = np.concatenate((temp_comb1, temp_comb2))
        random.shuffle(trialtypes)
        print(trialtypes)

        # counters for each trial type
        self.counter = np.zeros(4)


        #duration set up
        self.duration_rec_off = 6.5
        self.duration_rec_on_before = 4 #change this to exponential decay
        self.duration_odor_to_outcome = 1.3
        self.duration_water_large = 0.2

        self.duration_rec_on_after = 8
        self.duration_odor_on = 0.5


        for t in range(0, numtrials):

            print('trial number: ', t)
            print()
            self.settings.in_trial.update_value(True)
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=101)
            self.check_licking_1spout(self.duration_rec_on_before)


#           main training program
            self.run_trial_type(trialtypes[t])





            self.check_licking_1spout(self.duration_rec_on_after)
            self.settings.in_trial.update_value(False)
            # self.settings.save_video.update_value(False):


            self.wb.save(self.events_filename)

            self.check_licking_1spout(self.duration_rec_off)

            if self.interrupt_measurement_called:
                # tell subtherad to stop
                self.interrupt_subthread.emit()
                break


        odors_cue.initiate()
        odors_cue.close()
        print('FINISHED ASSOCIATION TRAINING')


        if self.settings.save_video.value():
            self.recorder.close()




    def check_licking_1spout(self, interval):

        checkperiod = 0.02
        timeout = time.time() + interval

        right_lick_last = 0
        while time.time() < timeout:
            right_lick = self.lickR.read()

            if right_lick != right_lick_last:
                if right_lick:
                    self.settings.lick_status.update_value(11)
                    print('Lick')
                    d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
                    d = self.ws.cell(row=self.ws.max_row, column=2, value=11)
                else:
                    self.settings.lick_status.update_value(10)
                    d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
                    d = self.ws.cell(row=self.ws.max_row, column=2, value=10)
            else:
                self.settings.lick_status.update_value(0)

            right_lick_last = right_lick
            time.sleep(checkperiod)

    def run_trial_type(self,types):
        odor_on = False
        reward_on = False
        if types == 0:
            print('contingency reward trial ' + str(self.counter[types]))
            print('opening odor port')
            odor_on = True
            reward_on = True
            r_code = [131, 130]
            w_code = [51, 50]
            self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_to_outcome)  ### can be a problem
            self.run_reward_module(reward_on, w_code)

        elif types == 1:
            print('contingency no reward trial ' + str(self.counter[types]))
            print('opening odor port')
            odor_on = True
            r_code = [141, 140]
            w_code = [61, 60]
            self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(duration_odor_to_outcome)  ### can be a problem
            self.run_reward_module(reward_on, w_code)
        elif types == 2:
            print('non-contingency reward trial ' + str(self.counter[types]))
            reward_on = True
            r_code = [151, 150]
            w_code = [71, 70]
            self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_to_outcome)  ### can be a problem
            self.run_reward_module(reward_on, w_code)
        else:
            print('non-contingency no reward trial ' + str(self.counter[types]))
            r_code = [161, 160]
            w_code = [71, 70]
            self.run_odor_module(odor_on, r_code)
            self.check_licking_1spout(self.duration_odor_to_outcome)  ### can be a problem
            self.run_reward_module(reward_on, w_code)

        self.counter[types] += 1

        self.wb.save(self.events_filename)




    def run_odor_module(self,odor_on, r_code):
        if odor_on:
            self.reward_odor.high()
            # self.OdorOnCopy.high()  # ？？？
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=r_code[0])

            time.sleep(self.duration_odor_on)

            print('closing odor port')
            self.reward_odor.low()
            self.OdorOnCopy.low()
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=r_code[1])
        else:
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=r_code[0])
            time.sleep(duration_odor_on)
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=r_code[1])

    def run_reward_module(self,reward_on, w_code):
        if reward_on:

            # modify! give water if licks three times within 1 s

            print('opening water valve')
            self.waterR.high()
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=w_code[0])
            self.check_licking_1spout(self.duration_water_large)  # this parameter hasn't een defined

            print('closing water valve')
            self.waterR.low()
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=w_code[1])

        else:
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=w_code[0])
            time.sleep(self.duration_water_large)
            d = self.ws.cell(row=(self.ws.max_row + 1), column=1, value=time.clock())
            d = self.ws.cell(row=self.ws.max_row, column=2, value=w_code[1])

#     def camera_action(self):
#         '''
#         format the image properly
#         '''
#         try:
#             wide_image = self.wide_cam.read()
#             wide_image_data = self.wide_cam.to_numpy(wide_image)
#             self.wide_disp_queue.put(wide_image_data)
#
#             if self.settings.save_video.value() and self.settings.in_trial.value() :
#                 self.recorder.save_frame(self.settings.filename.value(),wide_image)
#
#         except Exception as ex:
#             print('Error : %s' % ex)
class OdorGen(object):
    def __init__(self,odorindex):
        self.odorindex = odorindex
        self.odors_DAQ = []


    def assign_odor(self):
    # initiate all the odor solenoids

        for item in self.odorindex:

            self.odors_DAQ.append(DAQSimpleDOTask('Dev2_SELECT/port0/line{}'.format(item)))
        print('Odor {} has been properly assigned'.format(self.odorindex))


    def set_rewardodor(self,index):
        reward_odor = self.odors_DAQ[index]
        print('reward odor is odor {}'.format(index))
        return reward_odor


    def initiate(self):
        for odor in self.odors_DAQ:
            odor.low()
        print('Odor initiation: status low')
    def close(self):
        for odor in self.odors_DAQ:
            odor.close()
        print('Connection has been closed')

